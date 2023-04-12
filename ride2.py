import pandas as pd
import random
from datetime import datetime
import heapq
import calendar
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import time


file_path = "keskused.xlsx"

df = pd.read_excel(file_path, engine='openpyxl')

data_dict = {}

for index, row in df.iterrows():
    data_dict[index] = row.to_dict()

first_point = 'Endla 47, 10615, Tallinn'
min_length = 1117
max_length = 1190

mileages = [] # set of 12 mileages that
start_number = 314092

visited_locations = set()

def create_mileage():
    global mileage
    while True:
        row_number = random.choice(list(data_dict.keys()))
        if row_number not in visited_locations:
            center_data = data_dict[row_number]
            added_mileage = 2 * center_data['Kaugus aadressilt (km)']
            visited_locations.add(row_number)
            return added_mileage, center_data

def generate_monthly_excel(year, month, month_name, daily_data):
    wb = Workbook()
    ws = wb.active
    ws.title = month_name

    # Write header row
    header_row = ['Day', 'Route', 'Mileage', 'Total Mileage']
    for col_num, value in enumerate(header_row, 1):
        ws.cell(row=1, column=col_num, value=value)

    # Write daily data
    for row_num, row_data in enumerate(daily_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save the file in the new directory
    output_filename = f"{year}_{month:02d}_{month_name}.xlsx"
    output_directory = os.path.join("current_time", time.strftime("%M_%S"))
    os.makedirs(output_directory, exist_ok=True)
    wb.save(os.path.join(output_directory, output_filename))

def generate_calendar_year(year):
    month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    print(f"Year: {year}")
    print("--------------------------------------------------")
    annual_mileages = []

    for month_index, month_name in enumerate(month_names):
        print("--------------------------------------------------")  # Add this line to print dashes between month changes
        days_in_month = calendar.monthrange(year, month_index + 1)[1]  # Get the number of days in the current month

        daily_data = []
        for day in range(1, days_in_month + 1):
            print(f"{month_name} {day}:")
            daily_mileage = do_work(month_name)
            annual_mileages.append(daily_mileage)

            route = ", ".join([data_dict[row_number]['Keskuse nimi'] for row_number in visited_locations])

            daily_data.append((day, route, daily_mileage, start_number + daily_mileage))

            #print("--------------------------------------------------")

        generate_monthly_excel(year, month_index + 1, month_name, daily_data)

def do_work(month_name):
    global mileage, visited_locations
    mileage = 0 # Reset the mileage variable for each iteration
    candidates = []

    for _ in range(10):  # Generate 10 candidates
        mileage_candidate = 0
        visited_locations_candidate = set()

        while mileage_candidate < min_length:
            row_number = random.choice(list(data_dict.keys()))
            if row_number not in visited_locations_candidate:
                center_data = data_dict[row_number]
                added_mileage = 2 * center_data['Kaugus aadressilt (km)']
                visited_locations_candidate.add(row_number)
                if (mileage_candidate + added_mileage) <= max_length:
                    mileage_candidate += added_mileage
                else:
                    break
        # Add the candidate mileage and visited_locations to the candidates list
        candidates.append((mileage_candidate, visited_locations_candidate))

    # Select the smallest mileage variant over min_length or the second best
    candidates.sort(key=lambda x: x[0])  # Sort candidates by mileage
    best_candidate = None
    for candidate in candidates:
        if candidate[0] >= min_length:
            best_candidate = candidate
            break

    if best_candidate is None:
        # If no candidate meets the condition, select the second best
        best_candidate = heapq.nsmallest(2, candidates, key=lambda x: x[0])[-1]

    mileage, visited_locations = best_candidate

    return mileage  # Return the mileage

if __name__ == "__main__":
    generate_calendar_year(2022)